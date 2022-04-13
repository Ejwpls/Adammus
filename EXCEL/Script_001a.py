#Test out PYEXCEL

import openpyxl
path = r'C:\Users\edbert\Desktop\Python\EXCEL\test.xlsx'

#set variable
wb = openpyxl.load_workbook(path)

#test 01
print(type(wb))

#workbook sheet names
wb.sheetnames

print (wb.sheetnames[1])

#Get Drift Sheet
ActiveSheet = wb['Diaphragm Max Over Avg Drifts']
